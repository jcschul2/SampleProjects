from tkinter.filedialog import askdirectory

import openpyxl as py
import os
import csv

from openpyxl import Workbook


#  Extract gas chromatograph data in given directory, identify compound peaks based on retention time from list
#  provided (retention_times.xlsx), apply provided calibration curve, and output result in Excel file

def cvt_csv_to_xlsx(src_file_path, dst_file_path):  # Convert GC .csv output to .xlsx for compatibility with openpyxl
    wb = Workbook()
    ws = wb.active
    with open(src_file_path, 'r', encoding='utf8') as f:
        for row_csv in csv.reader(f):
            ws.append(row_csv)
    wb.save(dst_file_path)


path = askdirectory(title='Select Folder')
os.chdir(path)
runs = []  # List of each injection in the GC sequence

for item in os.listdir():  # Filter out non-GC run objects in directory
    if item[len(item) - 1] == 'D':
        runs.append(item)

peaks = Workbook()
peaks_sheet = peaks.active

RET = py.load_workbook('retention_times.xlsx')
RETs = RET.active
times = []

for time in RETs:  # List of retention time, calibration slope, and calibration y-intercept for each compound
    times.append([time[1].value, time[2].value, time[3].value])

areas = []  # List of lists of areas for each compound of interest for each sample

for sample in runs:  # Check each filtered run folder's report
    os.chdir(path)
    os.chdir(sample)
    try:
        cvt_csv_to_xlsx(os.listdir()[0], 'REPORT01.xlsx')  # GC reports are all called REPORT01.xlsx
    except FileNotFoundError:
        continue
    wb = py.load_workbook('REPORT01.xlsx')
    ws1 = wb["Sheet"]
    newAreasRow = list()  # List of peak areas for the current injection to be appended to areas
    newAreasRow.append(ws1['B3'].value)  # Retrieve name of injection sample
    for x in range(0, len(times)):
        newAreasRow.append(0)  # If no peak is detected, the area will be set to 0

    for (idx, cmpd) in enumerate(ws1["A"][10:]):  # Check each peak in the run
        for i in range(1, len(times) + 1):  # Check each peak against each known retention time
            try:
                if abs(float(eval(cmpd.value)) - times[i - 1][0]) < 0.04:  # Allow peak offset of 0.04 min
                    newAreasRow[i] = eval(ws1["D"][idx + 10].value)  # Peak areas start at row 11 in the worksheet
            except TypeError:
                continue
            except IndexError:
                continue
    areas.append(newAreasRow)

output = []

for (idxr, run2) in enumerate(areas):  # Apply calibration curve to extracted peak areas for each sample
    newCalibratedRow = [areas[idxr][0]]
    for (idxc, cmpd2) in enumerate(run2[1:]):
        try:
            if areas[idxr][idxc + 1] != 0:
                newCalibratedRow.append(max(0, areas[idxr][idxc + 1] / areas[idxr][1] * times[idxc][1] + times[idxc][2]))
            else:
                newCalibratedRow.append(0)  # If area is 0, set calibrated titer to 0 instead of applying calibration
        except ZeroDivisionError:
            continue
    output.append(newCalibratedRow)

writefinal = Workbook()
finalsheet = writefinal.active
firstrow = ['']
for row in RETs['A']:
    firstrow.append(row.value)  # Add names of each compound to top of output sheet
finalsheet.append(firstrow)
for row2 in output:  # Add sample names and calibrated titers for all samples
    finalsheet.append(row2)
os.chdir(path)
name = input('Filename: ')
writefinal.save(name)
